"""
=============================================================================
  Google Maps B2B Lead Scraper
  Autor: Script gerado para prospecção de clientes
  Descrição: Extrai dados de estabelecimentos no Google Maps e identifica
             leads prioritários (sem website cadastrado).
=============================================================================

  INSTALAÇÃO DAS DEPENDÊNCIAS:
  pip install playwright pandas openpyxl
  playwright install chromium

=============================================================================
"""

import time
import re
import sys
import pandas as pd
from datetime import datetime
from urllib.parse import quote

# ----------------------------------------------------------------------------
# IMPORTAÇÃO DO PLAYWRIGHT
# O Playwright é preferível ao Selenium por ser mais estável com SPAs (Single
# Page Applications) como o Google Maps, que usa carregamento dinâmico pesado.
# ----------------------------------------------------------------------------
try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
except ImportError:
    print("❌ Playwright não encontrado. Execute: pip install playwright && playwright install chromium")
    sys.exit(1)


# ============================================================================
#  CONFIGURAÇÕES GLOBAIS
# ============================================================================

# Tempo (em ms) para esperar elementos carregarem — aumente em conexões lentas
TIMEOUT_PADRAO = 10_000       # 10 segundos

# Pausa entre ações para simular comportamento humano e evitar bloqueios
PAUSA_SCROLL   = 1.5          # segundos entre cada scroll da lista
PAUSA_CLIQUE   = 2.0          # segundos após clicar em um estabelecimento
PAUSA_DETALHE  = 1.0          # segundos para aguardar detalhes carregarem

# Número máximo de resultados a coletar (None = sem limite)
MAX_RESULTADOS = 100

# ============================================================================
#  SELETORES CSS / XPATH
#  Estes são os seletores mais críticos do script. O Google Maps atualiza
#  sua interface com frequência, portanto eles podem precisar de ajuste.
#  Para depurar: abra o Maps no Chrome > F12 > Inspecionar elementos.
# ============================================================================

# Seletor da lista de resultados na barra lateral esquerda
SEL_LISTA_RESULTADOS = 'div[role="feed"]'

# Cada item (card) de estabelecimento na lista lateral
SEL_ITEM_LISTA = 'div[role="feed"] > div > div[jsaction]'

# Link clicável dentro de cada card de resultado
SEL_LINK_ITEM = 'a[href*="/maps/place/"]'

# --- Seletores do painel de detalhes (após clicar no estabelecimento) ---

# Nome do estabelecimento (h1 principal do painel)
SEL_NOME = 'h1.DUwDvf'

# Endereço — botão com data-item-id contendo "address"
SEL_ENDERECO = 'button[data-item-id="address"]'

# Telefone — botão com data-item-id começando com "phone:"
SEL_TELEFONE = 'button[data-item-id^="phone:"]'

# Website — âncora (link) com data-item-id="authority"
SEL_WEBSITE = 'a[data-item-id="authority"]'

# URL atual do Maps (usada para extrair o link do estabelecimento)
# Obtida diretamente via page.url após navegação


def limpar_texto(texto: str) -> str:
    """Remove espaços extras e caracteres indesejados de uma string."""
    if not texto:
        return ""
    return re.sub(r'\s+', ' ', texto.strip())


def obter_link_maps(page) -> str:
    """
    Retorna a URL canônica do estabelecimento no Google Maps.
    A URL muda dinamicamente após clicar no card.
    """
    url = page.url
    # Encurta URLs muito longas mantendo apenas a parte relevante
    match = re.search(r'(https://www\.google\.com/maps/place/[^?]+)', url)
    return match.group(1) if match else url


def extrair_detalhes(page) -> dict:
    """
    Extrai todos os dados de um estabelecimento já aberto no painel lateral.
    Usa try-except em cada campo para não interromper o scraping caso
    algum dado esteja ausente no perfil.

    Retorna um dicionário com os campos coletados.
    """
    dados = {
        "Nome":          "",
        "Endereço":      "",
        "Telefone":      "",
        "Website":       "",
        "Status do Site": "",
        "Link do Maps":  "",
    }

    # -- Nome ------------------------------------------------------------------
    try:
        elemento_nome = page.wait_for_selector(SEL_NOME, timeout=TIMEOUT_PADRAO)
        dados["Nome"] = limpar_texto(elemento_nome.inner_text())
    except PlaywrightTimeoutError:
        print("    ⚠️  Nome não encontrado — pulando estabelecimento.")
        return dados  # Sem nome, não vale a pena continuar

    # -- Endereço --------------------------------------------------------------
    try:
        el_end = page.query_selector(SEL_ENDERECO)
        if el_end:
            # O endereço fica no segundo filho de texto do botão
            dados["Endereço"] = limpar_texto(el_end.get_attribute("aria-label") or el_end.inner_text())
            # Limpa prefixo "Endereço: " se presente
            dados["Endereço"] = re.sub(r'^Endereço:\s*', '', dados["Endereço"])
    except Exception:
        pass  # Campo opcional

    # -- Telefone --------------------------------------------------------------
    try:
        el_tel = page.query_selector(SEL_TELEFONE)
        if el_tel:
            aria = el_tel.get_attribute("aria-label") or ""
            # Extrai apenas os dígitos e símbolos do número
            dados["Telefone"] = re.sub(r'^Telefone:\s*', '', aria).strip()
    except Exception:
        pass  # Campo opcional

    # -- Website ---------------------------------------------------------------
    try:
        el_site = page.query_selector(SEL_WEBSITE)
        if el_site:
            href = el_site.get_attribute("href") or ""
            # O Google usa redirecionamento — extrai a URL real do parâmetro q= ou url=
            match_url = re.search(r'[?&](?:q|url)=([^&]+)', href)
            if match_url:
                from urllib.parse import unquote
                dados["Website"] = unquote(match_url.group(1))
            else:
                dados["Website"] = href
    except Exception:
        pass  # Campo opcional

    # -- Status do Site --------------------------------------------------------
    dados["Status do Site"] = "Não Tem" if not dados["Website"] else "Tem"

    # -- Link do Maps ----------------------------------------------------------
    dados["Link do Maps"] = obter_link_maps(page)

    return dados


def scroll_lista(page, n_scrolls: int = 10) -> None:
    """
    Realiza scroll automático no feed lateral de resultados para carregar
    mais estabelecimentos. O Google Maps usa lazy-loading — novos cards
    só aparecem conforme o usuário rola a lista.

    Args:
        page:      Instância da página Playwright.
        n_scrolls: Quantas vezes rolar antes de tentar coletar resultados.
    """
    try:
        feed = page.wait_for_selector(SEL_LISTA_RESULTADOS, timeout=TIMEOUT_PADRAO)
    except PlaywrightTimeoutError:
        print("  ⚠️  Feed de resultados não encontrado para scroll.")
        return

    for i in range(n_scrolls):
        # Rola o elemento feed até o final usando JavaScript
        page.evaluate(
            "(el) => el.scrollBy(0, el.scrollHeight)",
            feed
        )
        time.sleep(PAUSA_SCROLL)
        print(f"  📜 Scroll {i + 1}/{n_scrolls}...", end="\r")

    print()  # Quebra de linha após os scrolls


def coletar_links_visiveis(page) -> list[str]:
    """
    Varre todos os cards visíveis na lista lateral e retorna
    uma lista de URLs de estabelecimentos para visitar.
    """
    links = page.query_selector_all(f'{SEL_ITEM_LISTA} {SEL_LINK_ITEM}')
    urls = []
    for link in links:
        href = link.get_attribute("href")
        if href and "/maps/place/" in href:
            # Normaliza a URL removendo parâmetros desnecessários
            url_limpa = href.split("?")[0]
            if url_limpa not in urls:
                urls.append(url_limpa)
    return urls


def scrape_google_maps(nicho: str, cidade: str, headless: bool = True) -> pd.DataFrame:
    """
    Função principal de scraping.

    Args:
        nicho:    Tipo de negócio (ex: "Dentistas", "Restaurantes").
        cidade:   Cidade ou região (ex: "São Paulo", "Curitiba PR").
        headless: Se True, o navegador roda invisível (sem janela gráfica).

    Returns:
        DataFrame do Pandas com todos os leads coletados.
    """
    # Monta a query de busca no formato que o Google Maps aceita via URL
    query = f"{nicho} em {cidade}"
    url_busca = f"https://www.google.com/maps/search/{quote(query)}"

    resultados = []

    print(f"\n🚀 Iniciando scraping: '{query}'")
    print(f"🌐 URL de busca: {url_busca}")
    print(f"🖥️  Modo headless: {'Sim' if headless else 'Não (janela visível)'}\n")

    with sync_playwright() as p:

        # ----------------------------------------------------------------
        # CONFIGURAÇÃO DO NAVEGADOR
        # Usamos Chromium por ser o mais compatível com o Google Maps.
        # slow_mo adiciona um delay (ms) entre cada ação — útil para debug.
        # ----------------------------------------------------------------
        browser = p.chromium.launch(
            headless=headless,
            slow_mo=50 if not headless else 0,  # Mais lento no modo visual
            args=[
                "--no-sandbox",
                "--disable-blink-features=AutomationControlled",  # Evita detecção de bot
                "--disable-dev-shm-usage",
            ]
        )

        # Cria contexto com configurações que imitam um usuário real
        context = browser.new_context(
            viewport={"width": 1366, "height": 768},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/121.0.0.0 Safari/537.36"
            ),
            locale="pt-BR",           # Garante interface em português
            timezone_id="America/Sao_Paulo",
        )

        page = context.new_page()

        # ----------------------------------------------------------------
        # NAVEGAÇÃO INICIAL
        # ----------------------------------------------------------------
        print("📍 Abrindo Google Maps...")
        page.goto(url_busca, wait_until="networkidle", timeout=30_000)
        time.sleep(3)

        # Fecha banner de cookies, se aparecer
        try:
            botao_aceitar = page.query_selector('button[aria-label*="Aceitar"]')
            if botao_aceitar:
                botao_aceitar.click()
                time.sleep(1)
        except Exception:
            pass

        # ----------------------------------------------------------------
        # COLETA DE LINKS — Scroll + Extração de URLs
        # ----------------------------------------------------------------
        print("📋 Carregando lista de resultados (aguarde os scrolls)...")

        todos_links = set()
        rodadas_sem_novos = 0
        MAX_RODADAS_SEM_NOVOS = 3  # Para de rolar se não aparecerem novos resultados

        while True:
            # Rola o feed para carregar mais
            scroll_lista(page, n_scrolls=5)

            # Coleta os links visíveis após o scroll
            links_atuais = coletar_links_visiveis(page)
            novos = set(links_atuais) - todos_links

            if not novos:
                rodadas_sem_novos += 1
                if rodadas_sem_novos >= MAX_RODADAS_SEM_NOVOS:
                    print(f"\n  ✅ Sem novos resultados após {MAX_RODADAS_SEM_NOVOS} tentativas. Lista completa.")
                    break
            else:
                rodadas_sem_novos = 0
                todos_links.update(novos)
                print(f"  🔗 {len(todos_links)} links coletados até agora...")

            # Respeita o limite máximo configurado
            if MAX_RESULTADOS and len(todos_links) >= MAX_RESULTADOS:
                print(f"\n  ⚠️  Limite de {MAX_RESULTADOS} resultados atingido.")
                break

        lista_links = list(todos_links)[:MAX_RESULTADOS]
        print(f"\n📊 Total de estabelecimentos para visitar: {len(lista_links)}\n")

        # ----------------------------------------------------------------
        # EXTRAÇÃO DE DETALHES — Visita cada link individualmente
        # ----------------------------------------------------------------
        for idx, link in enumerate(lista_links, start=1):
            print(f"  [{idx:03}/{len(lista_links)}] Extraindo: {link}")

            try:
                # Navega para a página do estabelecimento
                page.goto(link, wait_until="domcontentloaded", timeout=20_000)
                time.sleep(PAUSA_DETALHE)

                # Aguarda o painel de detalhes carregar
                page.wait_for_selector(SEL_NOME, timeout=TIMEOUT_PADRAO)
                time.sleep(PAUSA_CLIQUE)

                # Extrai os dados
                dados = extrair_detalhes(page)

                if dados["Nome"]:  # Só adiciona se tiver nome
                    resultados.append(dados)
                    status_emoji = "🔴" if dados["Status do Site"] == "Não Tem" else "🟢"
                    print(f"         ✔ {dados['Nome']} | Tel: {dados['Telefone'] or 'N/A'} | Site: {status_emoji} {dados['Status do Site']}")

            except PlaywrightTimeoutError:
                print(f"         ⏰ Timeout ao carregar — pulando.")
            except Exception as e:
                print(f"         ❌ Erro inesperado: {e} — pulando.")

            # Pequena pausa entre requests para não sobrecarregar o servidor
            time.sleep(0.5)

        browser.close()

    # ----------------------------------------------------------------
    # MONTA O DATAFRAME FINAL
    # ----------------------------------------------------------------
    if not resultados:
        print("\n⚠️  Nenhum resultado foi coletado.")
        return pd.DataFrame(columns=["Nome", "Status do Site", "Telefone", "Endereço", "Website", "Link do Maps"])

    df = pd.DataFrame(resultados)

    # Reordena colunas conforme especificado nos requisitos
    df = df[["Nome", "Status do Site", "Telefone", "Endereço", "Website", "Link do Maps"]]

    # Ordena: leads sem site primeiro (prioridade máxima de prospecção)
    df = df.sort_values(
        by="Status do Site",
        key=lambda x: x.map({"Não Tem": 0, "Tem": 1}),
        ascending=True
    ).reset_index(drop=True)

    return df


def salvar_resultados(df: pd.DataFrame, nicho: str, cidade: str) -> str:
    """
    Salva o DataFrame em arquivo .xlsx com formatação colorida.
    Leads sem site ficam destacados em vermelho claro.

    Returns:
        Caminho do arquivo salvo.
    """
    # Nome do arquivo com timestamp para evitar sobrescrever dados anteriores
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_arquivo = f"leads_{nicho.lower().replace(' ', '_')}_{cidade.lower().replace(' ', '_')}_{timestamp}.xlsx"
    caminho = nome_arquivo  # Salva no diretório atual

    try:
        with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Leads")

            workbook  = writer.book
            worksheet = writer.sheets["Leads"]

            # -- Formatação dos cabeçalhos --
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

            header_fill   = PatternFill("solid", fgColor="1F4E79")
            header_font   = Font(bold=True, color="FFFFFF", size=11)
            borda_fina    = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

            for cell in worksheet[1]:
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = borda_fina

            # -- Destaque de linhas por status do site --
            fill_sem_site = PatternFill("solid", fgColor="FFCCCC")  # Vermelho claro = lead quente
            fill_com_site = PatternFill("solid", fgColor="CCFFCC")  # Verde claro = já tem site

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                status_cell = row[1]  # Coluna B = "Status do Site"
                cor = fill_sem_site if status_cell.value == "Não Tem" else fill_com_site
                for cell in row:
                    cell.fill   = cor
                    cell.border = borda_fina
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

            # -- Ajusta largura das colunas automaticamente --
            larguras = {"A": 40, "B": 15, "C": 18, "D": 45, "E": 35, "F": 55}
            for col, largura in larguras.items():
                worksheet.column_dimensions[col].width = largura

            # -- Congela a primeira linha (cabeçalho) --
            worksheet.freeze_panes = "A2"

        print(f"\n✅ Arquivo salvo com sucesso: {caminho}")

    except Exception as e:
        # Fallback para CSV caso openpyxl falhe
        print(f"⚠️  Erro ao salvar .xlsx: {e}. Salvando como .csv...")
        caminho = caminho.replace(".xlsx", ".csv")
        df.to_csv(caminho, index=False, encoding="utf-8-sig")
        print(f"✅ Arquivo CSV salvo: {caminho}")

    return caminho


def exibir_resumo(df: pd.DataFrame, nicho: str, cidade: str) -> None:
    """Exibe um resumo dos resultados no terminal."""
    total       = len(df)
    sem_site    = len(df[df["Status do Site"] == "Não Tem"])
    com_site    = len(df[df["Status do Site"] == "Tem"])
    com_telefone = len(df[df["Telefone"] != ""])

    print("\n" + "=" * 55)
    print("  📊  RESUMO DA PROSPECÇÃO")
    print("=" * 55)
    print(f"  🔍  Busca:              {nicho} em {cidade}")
    print(f"  📋  Total de leads:     {total}")
    print(f"  🔴  Sem website:        {sem_site}  ← LEADS PRIORITÁRIOS")
    print(f"  🟢  Com website:        {com_site}")
    print(f"  📞  Com telefone:       {com_telefone}")
    print("=" * 55)

    if sem_site > 0:
        print(f"\n  💡  TOP 5 LEADS QUENTES (sem site):\n")
        top = df[df["Status do Site"] == "Não Tem"].head(5)
        for _, row in top.iterrows():
            print(f"     • {row['Nome']}")
            print(f"       📞 {row['Telefone'] or 'Telefone não disponível'}")
            print(f"       📍 {row['Endereço'] or 'Endereço não disponível'}\n")


# ============================================================================
#  GERADOR DE PROMPTS DE SITE
#  Para cada lead sem website, gera um prompt preenchido com os dados
#  coletados, pronto para colar no Claude e gerar uma landing page completa.
# ============================================================================

TEMPLATE_PROMPT = """Você é um desenvolvedor web senior e designer especialista em landing pages de alta conversão para prestadores de serviços locais no Brasil. Crie um site completo em HTML, CSS e JavaScript puro (arquivo único) seguindo TODAS as instruções abaixo com precisão.

---

### 🏢 DADOS DA EMPRESA

- **Nome da empresa:** {nome}
- **Segmento / tipo de serviço:** {segmento}
- **Cidade / região de atuação:** {cidade}
- **Número WhatsApp (com DDI):** {whatsapp}
- **Cor primária da marca (hex):** {cor_primaria}
- **Cor secundária / de apoio (hex):** {cor_secundaria}
- **URL da logo:** sem logo
- **Serviços oferecidos:** {servicos}
- **Diferenciais da empresa:** {diferenciais}
- **Depoimentos de clientes:** (gere 3 depoimentos fictícios verossímeis para o segmento)
- **CNPJ:** não informado
- **E-mail:** não informado
- **Redes sociais:** não informado
- **Funciona quantas horas / dias?** {horario}
- **Tom de comunicação desejado:** {tom}

---

### 🎨 REGRAS DE DESIGN (OBRIGATÓRIAS)

1. **DESIGN ÚNICO E INOVADOR**: O layout deve ser original e visualmente impactante. PROIBIDO usar estruturas genéricas. Escolha UMA direção estética forte e execute com precisão:
   - Brutalist / Raw (tipografia pesada, contraste extremo, formas cortadas)
   - Editorial Magazine (assimetria elegante, colunas criativas, espaços generosos)
   - Dark Luxury (fundo escuro, detalhes dourados ou vibrantes, sofisticação)
   - Industrial Bold (texturas de metal/concreto, tipografia condensada, badges técnicos)
   - Organic & Clean (gradientes suaves, formas arredondadas, leveza visual)
   - Neon Urban (fundo escuro + neon da cor da marca, energia urbana)
   - Minimalismo Geométrico (linhas precisas, grid rigoroso, espaço em branco proposital)

2. **TIPOGRAFIA DISTINTIVA**: Importe fontes do Google Fonts que combinem com o segmento. PROIBIDO usar Inter, Roboto ou Arial.

3. **PALETA DE CORES**: Use a cor primária da empresa como dominante. Verde WhatsApp (#25D366) APENAS nos botões de contato.

4. **ANIMAÇÕES FUNCIONAIS**: Inclua obrigatoriamente:
   - Pulsação no botão de WhatsApp principal
   - Fade-in + slide nos elementos ao rolar a página (IntersectionObserver)
   - Hover states expressivos nos cards e botões
   - Contador animado de números

5. **LAYOUT DINÂMICO**: Use pelo menos 2 técnicas de composição avançadas.

---

### 📐 ESTRUTURA OBRIGATÓRIA DO SITE

1. HEADER FIXO com logo, menu desktop e botão WhatsApp
2. HERO SECTION com H1 "{segmento} em {cidade}", 2 CTAs e badge de disponibilidade
3. SEÇÃO DE SERVIÇOS em grid de cards com botão WhatsApp específico por serviço
4. SEÇÃO "POR QUE NOS ESCOLHER" com diferenciais e contadores animados
5. DEPOIMENTOS com avatares, estrelas e avaliação média
6. SEÇÃO DE ORÇAMENTO RÁPIDO com urgência e botão WhatsApp gigante
7. FAQ com accordion (4-6 perguntas relevantes ao segmento, SEO semântico)
8. FOOTER completo
9. BOTÃO FLUTUANTE FIXO do WhatsApp com pulsação

---

### 📱 RESPONSIVIDADE, SEO E WHATSAPP

- Mobile-first, menu hamburger, grid adaptativo
- Meta tags SEO completas, H1 único, Schema.org LocalBusiness JSON-LD
- Todos os links WhatsApp com mensagens pré-preenchidas contextuais
- URL padrão: https://wa.me/{whatsapp_raw}?text=[mensagem codificada]

---

### ⚙️ TÉCNICO

- HTML5 semântico, Tailwind CSS via CDN, Google Fonts, JavaScript puro
- Animações com @keyframes e IntersectionObserver
- Cores como variáveis CSS no :root
- Arquivo único .html completo e funcional

**Gere o HTML completo, funcional e pronto para publicar. Não use comentários de TODO ou placeholders no código final.**"""


def inferir_dados_segmento(nome: str, nicho: str, cidade: str) -> dict:
    """
    Infere dados padrão de design e copy com base no nicho detectado.
    Retorna um dict com cor, serviços, diferenciais, horário e tom.
    """

    nicho_lower = nicho.lower()

    # --- Mapeamento de nichos para paletas e diferenciais padrão ---
    configs = {
        "dedetiz": {
            "cor_primaria": "#E21F26", "cor_secundaria": "#1a1a1a",
            "servicos": "Dedetização, Desratização, Descupinização, Sanitização, Controle de pombos, Desinsetização",
            "diferenciais": "Atendimento 24h, Produtos certificados ANVISA, Garantia de 90 dias, Equipe treinada, Orçamento grátis",
            "horario": "24h / 7 dias por semana", "tom": "urgente e direto",
        },
        "eletric": {
            "cor_primaria": "#F5A623", "cor_secundaria": "#0a0a0a",
            "servicos": "Instalações elétricas, Manutenção preventiva, Laudos elétricos, Quadros de distribuição, Iluminação",
            "diferenciais": "Eletricistas certificados NR10, Atendimento rápido, Garantia nos serviços, Orçamento no local, Equipe especializada",
            "horario": "24h / 7 dias por semana", "tom": "confiável e técnico",
        },
        "encanament": {
            "cor_primaria": "#1E90FF", "cor_secundaria": "#0a0a0a",
            "servicos": "Desentupimento, Vazamentos, Instalação hidráulica, Limpeza de caixas d'água, Conserto de torneiras",
            "diferenciais": "Atendimento emergencial, Equipe especializada, Garantia nos serviços, Orçamento grátis, Preço justo",
            "horario": "24h / 7 dias por semana", "tom": "urgente e direto",
        },
        "limpez": {
            "cor_primaria": "#00C49A", "cor_secundaria": "#f9f9f9",
            "servicos": "Limpeza residencial, Limpeza comercial, Limpeza pós-obra, Lavagem de sofás e tapetes, Higienização",
            "diferenciais": "Produtos ecológicos, Equipe uniformizada, Pontualidade garantida, Orçamento grátis, Seguro contra danos",
            "horario": "Segunda a sábado, 8h às 18h", "tom": "amigável e próximo",
        },
        "pintur": {
            "cor_primaria": "#FF6B35", "cor_secundaria": "#1a1a1a",
            "servicos": "Pintura residencial, Pintura comercial, Textura, Grafiato, Pintura externa e interna",
            "diferenciais": "Acabamento impecável, Materiais de qualidade, Prazo garantido, Orçamento grátis, Experiência comprovada",
            "horario": "Segunda a sábado, 7h às 17h", "tom": "confiável e direto",
        },
        "ar.condicion": {
            "cor_primaria": "#0099CC", "cor_secundaria": "#f0f0f0",
            "servicos": "Instalação de ar-condicionado, Manutenção preventiva, Higienização, Recarga de gás, Conserto",
            "diferenciais": "Técnicos certificados, Peças originais, Atendimento rápido, Garantia de serviço, Orçamento grátis",
            "horario": "Segunda a sábado, 8h às 18h", "tom": "técnico e confiável",
        },
        "dentist": {
            "cor_primaria": "#2E86AB", "cor_secundaria": "#f9f9f9",
            "servicos": "Clareamento dental, Ortodontia, Implantes, Limpeza, Restaurações, Próteses dentárias",
            "diferenciais": "Clínica moderna, Equipe especializada, Ambiente confortável, Parcelamento facilitado, Atendimento humanizado",
            "horario": "Segunda a sexta, 8h às 18h", "tom": "confiável e acolhedor",
        },
        "academi": {
            "cor_primaria": "#FF3D00", "cor_secundaria": "#0a0a0a",
            "servicos": "Musculação, Funcional, Spinning, Yoga, Personal trainer, Aulas em grupo",
            "diferenciais": "Equipamentos modernos, Professores qualificados, Ambiente climatizado, Planos flexíveis, Avaliação física gratuita",
            "horario": "Segunda a sábado, 6h às 22h", "tom": "energético e motivador",
        },
    }

    # Detecta o nicho pelo nome
    config = None
    for chave, dados in configs.items():
        if re.search(chave, nicho_lower):
            config = dados
            break

    # Fallback genérico se nicho não mapeado
    if not config:
        config = {
            "cor_primaria": "#2563EB", "cor_secundaria": "#1a1a1a",
            "servicos": f"Serviços de {nicho} em {cidade}",
            "diferenciais": "Atendimento profissional, Qualidade garantida, Orçamento grátis, Equipe qualificada, Preço justo",
            "horario": "Segunda a sábado, 8h às 18h", "tom": "profissional e direto",
        }

    return config


def gerar_prompt_site(lead: dict, nicho: str, cidade: str) -> str:
    """
    Recebe os dados de um lead (sem site) e retorna o prompt completo
    preenchido, pronto para ser colado no Claude e gerar a landing page.

    Args:
        lead:   Dicionário com Nome, Telefone, Endereço, etc.
        nicho:  Nicho da busca (ex: "Dentistas")
        cidade: Cidade da busca (ex: "São Paulo SP")
    """
    config = inferir_dados_segmento(lead["Nome"], nicho, cidade)

    # Limpa o telefone para uso na URL do WhatsApp (somente dígitos)
    telefone_raw = re.sub(r'\D', '', lead.get("Telefone", ""))
    # Adiciona DDI 55 se não tiver
    if telefone_raw and not telefone_raw.startswith("55"):
        telefone_raw = "55" + telefone_raw

    whatsapp_display = lead.get("Telefone") or telefone_raw or "não informado"

    prompt = TEMPLATE_PROMPT.format(
        nome          = lead["Nome"],
        segmento      = nicho,
        cidade        = cidade,
        whatsapp      = whatsapp_display,
        whatsapp_raw  = telefone_raw or "5500000000000",
        cor_primaria  = config["cor_primaria"],
        cor_secundaria= config["cor_secundaria"],
        servicos      = config["servicos"],
        diferenciais  = config["diferenciais"],
        horario       = config["horario"],
        tom           = config["tom"],
    )

    return prompt


def salvar_prompts(df: pd.DataFrame, nicho: str, cidade: str) -> str:
    """
    Para cada lead sem site, gera e salva o prompt em um arquivo .txt
    com separadores claros. Retorna o caminho do arquivo gerado.
    """
    leads_sem_site = df[df["Status do Site"] == "Não Tem"]

    if leads_sem_site.empty:
        return None

    timestamp    = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_arquivo = f"prompts_sites_{nicho.lower().replace(' ', '_')}_{cidade.lower().replace(' ', '_')}_{timestamp}.txt"

    linhas = []
    linhas.append("=" * 80)
    linhas.append(f"  PROMPTS PARA GERAÇÃO DE SITES — {nicho.upper()} EM {cidade.upper()}")
    linhas.append(f"  Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    linhas.append(f"  Total de prompts: {len(leads_sem_site)}")
    linhas.append("=" * 80)
    linhas.append("")
    linhas.append("  COMO USAR:")
    linhas.append("  1. Copie o bloco de prompt desejado (entre as linhas ▼ e ▲)")
    linhas.append("  2. Cole no Claude (claude.ai) ou em qualquer LLM de sua preferência")
    linhas.append("  3. O site HTML completo será gerado automaticamente")
    linhas.append("  4. Salve o resultado como index.html e publique")
    linhas.append("")

    for idx, (_, row) in enumerate(leads_sem_site.iterrows(), start=1):
        lead = row.to_dict()

        linhas.append("─" * 80)
        linhas.append(f"  LEAD #{idx:02} — {lead['Nome']}")
        linhas.append(f"  📞 {lead['Telefone'] or 'Telefone não informado'}")
        linhas.append(f"  📍 {lead['Endereço'] or 'Endereço não informado'}")
        linhas.append(f"  🔗 {lead['Link do Maps']}")
        linhas.append("─" * 80)
        linhas.append("")
        linhas.append("  ▼ COPIE O PROMPT ABAIXO ATÉ A LINHA ▲ ▼ ▼ ▼ ▼ ▼ ▼ ▼ ▼ ▼ ▼ ▼")
        linhas.append("")

        prompt = gerar_prompt_site(lead, nicho, cidade)
        linhas.append(prompt)

        linhas.append("")
        linhas.append("  ▲ FIM DO PROMPT ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲ ▲")
        linhas.append("")

    with open(nome_arquivo, "w", encoding="utf-8") as f:
        f.write("\n".join(linhas))

    return nome_arquivo


# ============================================================================
#  PONTO DE ENTRADA
# ============================================================================

if __name__ == "__main__":

    print("=" * 55)
    print("  🗺️   GOOGLE MAPS B2B LEAD SCRAPER")
    print("=" * 55)

    # -- Entrada do usuário --------------------------------------------------
    nicho  = input("\n  📌 Digite o nicho de mercado (ex: Dentistas): ").strip()
    cidade = input("  🏙️  Digite a cidade/região (ex: São Paulo SP): ").strip()

    if not nicho or not cidade:
        print("❌ Nicho e cidade são obrigatórios!")
        sys.exit(1)

    # Pergunta sobre modo headless
    resposta_headless = input("\n  🖥️  Rodar em modo invisível/headless? (S/N) [padrão: S]: ").strip().upper()
    headless = resposta_headless != "N"

    # -- Execução do scraping ------------------------------------------------
    try:
        df_leads = scrape_google_maps(nicho, cidade, headless=headless)

        if not df_leads.empty:
            # Salva a planilha de resultados
            arquivo_xlsx = salvar_resultados(df_leads, nicho, cidade)

            # Exibe resumo no terminal
            exibir_resumo(df_leads, nicho, cidade)

            # Gera os prompts de site para todos os leads sem website
            sem_site = len(df_leads[df_leads["Status do Site"] == "Não Tem"])
            if sem_site > 0:
                print(f"\n  🤖 Gerando prompts de site para {sem_site} lead(s) sem website...")
                arquivo_prompts = salvar_prompts(df_leads, nicho, cidade)
                if arquivo_prompts:
                    print(f"  ✅ Prompts salvos em: {arquivo_prompts}")
                    print(f"  💡 Abra o arquivo, copie qualquer bloco de prompt e cole no Claude!")

            print(f"\n  📁 Planilha gerada: {arquivo_xlsx}\n")
        else:
            print("\n❌ Nenhum dado coletado. Verifique a busca e tente novamente.")

    except KeyboardInterrupt:
        print("\n\n⚠️  Scraping interrompido pelo usuário.")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Erro fatal: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)