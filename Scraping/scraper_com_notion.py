"""
=============================================================================
  Google Maps B2B Lead Scraper  ·  com integração Notion
  Autor: Script gerado para prospecção de clientes
  Descrição: Extrai dados de estabelecimentos no Google Maps, identifica
             leads prioritários (sem website) e os envia automaticamente
             para o Notion.

  NOVIDADE: ao final do scraping, todos os leads sem site são enviados
  direto para sua Base de Leads no Notion — sem copiar e colar nada.

=============================================================================

  INSTALAÇÃO DAS DEPENDÊNCIAS:
  pip install playwright pandas openpyxl requests python-dotenv
  playwright install chromium

  CONFIGURAÇÃO DO NOTION (uma vez só):
  1. Crie o arquivo .env na mesma pasta deste script com:
       NOTION_TOKEN=secret_xxxxxxxxxxxx
       DATABASE_ID=xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx
  2. Siga as instruções em notion_sync.py para obter esses valores.

=============================================================================
"""

import time
import re
import sys
import pandas as pd
from datetime import datetime
from urllib.parse import quote

# ── Integração com Notion ────────────────────────────────────────────────────
# Importa o módulo de integração. Se não encontrar, desativa silenciosamente.
try:
    from notion_sync import enviar_lote_para_notion
    NOTION_DISPONIVEL = True
except ImportError:
    NOTION_DISPONIVEL = False
    print("⚠️  notion_sync.py não encontrado — scraping funcionará normalmente,")
    print("   mas os leads NÃO serão enviados ao Notion.\n")

# ── Playwright ───────────────────────────────────────────────────────────────
try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
except ImportError:
    print("❌ Playwright não encontrado. Execute: pip install playwright && playwright install chromium")
    sys.exit(1)


# ============================================================================
#  CONFIGURAÇÕES GLOBAIS
# ============================================================================

TIMEOUT_PADRAO = 8_000
PAUSA_SCROLL   = 0.8
PAUSA_CLIQUE   = 1.0
PAUSA_DETALHE  = 0.5
MAX_RESULTADOS = 100

# ============================================================================
#  SELETORES CSS
# ============================================================================

SEL_LISTA_RESULTADOS = 'div[role="feed"]'
SEL_ITEM_LISTA       = 'div[role="feed"] > div > div[jsaction]'
SEL_LINK_ITEM        = 'a[href*="/maps/place/"]'
SEL_NOME             = 'h1.DUwDvf'
SEL_ENDERECO         = 'button[data-item-id="address"]'
SEL_TELEFONE         = 'button[data-item-id^="phone:"]'
SEL_WEBSITE          = 'a[data-item-id="authority"]'


def limpar_texto(texto: str) -> str:
    if not texto:
        return ""
    return re.sub(r'\s+', ' ', texto.strip())


def obter_link_maps(page) -> str:
    url   = page.url
    match = re.search(r'(https://www\.google\.com/maps/place/[^?]+)', url)
    return match.group(1) if match else url


def extrair_detalhes(page) -> dict:
    dados = {
        "Nome":           "",
        "Endereço":       "",
        "Telefone":       "",
        "Website":        "",
        "Status do Site": "",
        "Link do Maps":   "",
    }

    try:
        elemento_nome  = page.wait_for_selector(SEL_NOME, timeout=TIMEOUT_PADRAO)
        dados["Nome"]  = limpar_texto(elemento_nome.inner_text())
    except PlaywrightTimeoutError:
        return dados

    try:
        el_end = page.query_selector(SEL_ENDERECO)
        if el_end:
            dados["Endereço"] = limpar_texto(el_end.get_attribute("aria-label") or el_end.inner_text())
            dados["Endereço"] = re.sub(r'^Endereço:\s*', '', dados["Endereço"])
    except Exception:
        pass

    try:
        el_tel = page.query_selector(SEL_TELEFONE)
        if el_tel:
            aria = el_tel.get_attribute("aria-label") or ""
            dados["Telefone"] = re.sub(r'^Telefone:\s*', '', aria).strip()
    except Exception:
        pass

    try:
        el_site = page.query_selector(SEL_WEBSITE)
        if el_site:
            href      = el_site.get_attribute("href") or ""
            match_url = re.search(r'[?&](?:q|url)=([^&]+)', href)
            if match_url:
                from urllib.parse import unquote
                dados["Website"] = unquote(match_url.group(1))
            else:
                dados["Website"] = href
    except Exception:
        pass

    dados["Status do Site"] = "Não Tem" if not dados["Website"] else "Tem"
    dados["Link do Maps"]   = obter_link_maps(page)

    return dados


def scroll_lista(page, n_scrolls: int = 10) -> None:
    try:
        feed = page.wait_for_selector(SEL_LISTA_RESULTADOS, timeout=TIMEOUT_PADRAO)
    except PlaywrightTimeoutError:
        return

    for i in range(n_scrolls):
        page.evaluate("(el) => el.scrollBy(0, el.scrollHeight)", feed)
        time.sleep(PAUSA_SCROLL)


def coletar_links_visiveis(page) -> list[str]:
    links = page.query_selector_all(f'{SEL_ITEM_LISTA} {SEL_LINK_ITEM}')
    urls  = []
    for link in links:
        href = link.get_attribute("href")
        if href and "/maps/place/" in href:
            url_limpa = href.split("?")[0]
            if url_limpa not in urls:
                urls.append(url_limpa)
    return urls


def scrape_google_maps(nicho: str, cidade: str, headless: bool = True) -> pd.DataFrame:
    query     = f"{nicho} em {cidade}"
    url_busca = f"https://www.google.com/maps/search/{quote(query)}"
    resultados = []

    print(f"\n{'=' * 55}")
    print(f"  🗺️  INICIANDO SCRAPING")
    print(f"{'=' * 55}")
    print(f"  🔍  Busca:    {query}")
    print(f"  🖥️  Headless: {'Sim' if headless else 'Não'}")
    print(f"{'=' * 55}\n")

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=headless,
            slow_mo=50 if not headless else 0,
            args=[
                "--no-sandbox",
                "--disable-blink-features=AutomationControlled",
                "--disable-dev-shm-usage",
            ]
        )
        context = browser.new_context(
            viewport={"width": 1366, "height": 768},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/121.0.0.0 Safari/537.36"
            ),
            locale="pt-BR",
            timezone_id="America/Sao_Paulo",
        )
        page = context.new_page()

        page.goto(url_busca, wait_until="domcontentloaded", timeout=30_000)
        time.sleep(3)

        try:
            botao_aceitar = page.query_selector('button[aria-label*="Aceitar"]')
            if botao_aceitar:
                botao_aceitar.click()
                time.sleep(1)
        except Exception:
            pass

        todos_links       = set()
        rodadas_sem_novos = 0
        MAX_RODADAS_SEM_NOVOS = 3

        while True:
            scroll_lista(page, n_scrolls=5)
            links_atuais = coletar_links_visiveis(page)
            novos        = set(links_atuais) - todos_links

            if not novos:
                rodadas_sem_novos += 1
                if rodadas_sem_novos >= MAX_RODADAS_SEM_NOVOS:
                    break
            else:
                rodadas_sem_novos = 0
                todos_links.update(novos)

            if MAX_RESULTADOS and len(todos_links) >= MAX_RESULTADOS:
                break

        lista_links = list(todos_links)[:MAX_RESULTADOS]
        print(f"  📋 {len(lista_links)} estabelecimentos encontrados. Extraindo dados...\n")

        for idx, link in enumerate(lista_links, start=1):
            # Progresso em linha única que se atualiza
            print(f"  ⏳ Progresso: {idx}/{len(lista_links)}  ({int(idx/len(lista_links)*100)}%)", end="\r")
            try:
                page.goto(link, wait_until="domcontentloaded", timeout=20_000)
                time.sleep(PAUSA_DETALHE)
                page.wait_for_selector(SEL_NOME, timeout=TIMEOUT_PADRAO)
                time.sleep(PAUSA_CLIQUE)

                dados = extrair_detalhes(page)
                if dados["Nome"]:
                    resultados.append(dados)

            except PlaywrightTimeoutError:
                pass
            except Exception:
                pass

            time.sleep(0.2)

        print(f"  ✅ Extração concluída: {len(resultados)} leads coletados.          \n")
        browser.close()

    if not resultados:
        print("\n⚠️  Nenhum resultado foi coletado.")
        return pd.DataFrame(columns=["Nome", "Status do Site", "Telefone", "Endereço", "Website", "Link do Maps"])

    df = pd.DataFrame(resultados)
    df = df[["Nome", "Status do Site", "Telefone", "Endereço", "Website", "Link do Maps"]]
    df = df.sort_values(
        by="Status do Site",
        key=lambda x: x.map({"Não Tem": 0, "Tem": 1}),
        ascending=True
    ).reset_index(drop=True)

    return df


def salvar_resultados(df: pd.DataFrame, nicho: str, cidade: str) -> str:
    timestamp    = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_arquivo = f"leads_{nicho.lower().replace(' ', '_')}_{cidade.lower().replace(' ', '_')}_{timestamp}.xlsx"
    caminho      = nome_arquivo

    try:
        with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Leads")
            workbook  = writer.book
            worksheet = writer.sheets["Leads"]

            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            header_fill = PatternFill("solid", fgColor="1F4E79")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            borda_fina  = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"),  bottom=Side(style="thin")
            )
            for cell in worksheet[1]:
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = borda_fina

            fill_sem_site = PatternFill("solid", fgColor="FFCCCC")
            fill_com_site = PatternFill("solid", fgColor="CCFFCC")
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                status_cell = row[1]
                cor = fill_sem_site if status_cell.value == "Não Tem" else fill_com_site
                for cell in row:
                    cell.fill      = cor
                    cell.border    = borda_fina
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

            larguras = {"A": 40, "B": 15, "C": 18, "D": 45, "E": 35, "F": 55}
            for col, largura in larguras.items():
                worksheet.column_dimensions[col].width = largura
            worksheet.freeze_panes = "A2"

        print(f"\n✅ Arquivo salvo com sucesso: {caminho}")
    except Exception as e:
        print(f"⚠️  Erro ao salvar .xlsx: {e}. Salvando como .csv...")
        caminho = caminho.replace(".xlsx", ".csv")
        df.to_csv(caminho, index=False, encoding="utf-8-sig")
        print(f"✅ Arquivo CSV salvo: {caminho}")

    return caminho


def exibir_resumo(df: pd.DataFrame, nicho: str, cidade: str) -> None:
    total        = len(df)
    sem_site     = len(df[df["Status do Site"] == "Não Tem"])
    com_site     = len(df[df["Status do Site"] == "Tem"])
    com_telefone = len(df[df["Telefone"] != ""])

    print("\n" + "=" * 55)
    print("  📊  RESUMO DA PROSPECÇÃO")
    print("=" * 55)
    print(f"  🔍  Busca:              {nicho} em {cidade}")
    print(f"  📋  Total de leads:     {total}")
    print(f"  🔴  Sem website:        {sem_site}  ← LEADS PRIORITÁRIOS")
    print(f"  🟢  Com website:        {com_site}")
    print(f"  📞  Com telefone:       {com_telefone}")
    print("=" * 55)

    if sem_site > 0:
        print(f"\n  💡  TOP 5 LEADS QUENTES (sem site):\n")
        top = df[df["Status do Site"] == "Não Tem"].head(5)
        for _, row in top.iterrows():
            print(f"     • {row['Nome']}")
            print(f"       📞 {row['Telefone'] or 'Telefone não disponível'}")
            print(f"       📍 {row['Endereço'] or 'Endereço não disponível'}\n")


# ── O restante do código (prompts, template, etc.) permanece igual ────────────
# (Cole aqui as funções gerar_prompt_site, salvar_prompts, inferir_dados_segmento
#  e TEMPLATE_PROMPT do seu arquivo original — não foram alteradas)


# ============================================================================
#  PONTO DE ENTRADA
# ============================================================================

if __name__ == "__main__":

    print("=" * 55)
    print("  🗺️   GOOGLE MAPS B2B LEAD SCRAPER  +  NOTION")
    print("=" * 55)

    nicho  = input("\n  📌 Digite o nicho de mercado (ex: Dentistas): ").strip()
    cidade = input("  🏙️  Digite a cidade/região (ex: São Paulo SP): ").strip()

    if not nicho or not cidade:
        print("❌ Nicho e cidade são obrigatórios!")
        sys.exit(1)

    resposta_headless = input("\n  🖥️  Rodar em modo invisível/headless? (S/N) [padrão: S]: ").strip().upper()
    headless = resposta_headless != "N"

    # ── Pergunta sobre o Notion ──────────────────────────────────────────────
    usar_notion = False
    rodada_nome = ""

    if NOTION_DISPONIVEL:
        resp_notion = input("\n  📋 Enviar leads sem site direto para o Notion? (S/N) [padrão: S]: ").strip().upper()
        usar_notion = resp_notion != "N"

        if usar_notion:
            rodada_nome = input("  📦 Nome da rodada/lote (ex: Lote 01 - Março) [Enter para deixar em branco]: ").strip()
            print()

    try:
        df_leads = scrape_google_maps(nicho, cidade, headless=headless)

        if not df_leads.empty:
            arquivo_xlsx = salvar_resultados(df_leads, nicho, cidade)
            exibir_resumo(df_leads, nicho, cidade)

            # ── ENVIO AUTOMÁTICO PARA O NOTION ──────────────────────────────
            if usar_notion and NOTION_DISPONIVEL:
                enviar_lote_para_notion(
                    df     = df_leads,
                    nicho  = nicho,
                    rodada = rodada_nome,
                )
            # ────────────────────────────────────────────────────────────────

            sem_site = len(df_leads[df_leads["Status do Site"] == "Não Tem"])
            if sem_site > 0:
                print(f"\n  🤖 Gerando prompts de site para {sem_site} lead(s) sem website...")
                # arquivo_prompts = salvar_prompts(df_leads, nicho, cidade)  # descomente se usar

            print(f"\n  📁 Planilha gerada: {arquivo_xlsx}\n")
        else:
            print("\n❌ Nenhum dado coletado. Verifique a busca e tente novamente.")

    except KeyboardInterrupt:
        print("\n\n⚠️  Scraping interrompido pelo usuário.")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Erro fatal: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)